import calendar
import os
import sys
import shutil
from datetime import datetime, timedelta
from dateutil import relativedelta
from openpyxl import load_workbook


def create_new_biweekly_finance_file(file_date: datetime, template_file_path: str):
    """
    Creates a new biweekly_finance file from the template and renames it based on it's date
    :param template_file_path: file path to the template biweekly finance .xlsx file
    :param file_date: month and year for the new biweekly finance file
    :return: filepath of the new biweekly finance file
    :rtype: str
    """
    new_file_path = _determine_biweekly_finance_file_path(file_date=file_date)
    new_biweekly_finance_file_dir = os.path.dirname(new_file_path)
    new_biweekly_finance_file_name = os.path.basename(new_file_path)
    new_file_path = _copy_and_rename_file(file_path=template_file_path,
                                          dest_dir=new_biweekly_finance_file_dir,
                                          new_file_name=new_biweekly_finance_file_name)
    return new_file_path


def populate_biweekly_finance_values(new_file_path: str, previous_file_path: str):
    """
    Populates all values of the biweekly finance file as a successor to the previous file
    :param new_file_path: new biweekly finance file to populate with values
    :param previous_file_path: file path of biweekly finance file of the previous month
    """
    # New File
    new_workbook = load_workbook(new_file_path)
    new_workbook_datetime = _determine_biweekly_finance_file_datetime(new_file_path)
    new_month_abbr = new_workbook_datetime.strftime('%b')

    # Rename New File Worksheet Names
    print("Renaming worksheets...")
    new_workbook["Pay 1"].title = "{0} - Pay 1".format(new_month_abbr)
    new_workbook["Pay 2"].title = "{0} - Pay 2".format(new_month_abbr)
    new_workbook["Pay 3"].title = "{0} - Pay 3".format(new_month_abbr)
    new_workbook.save(new_file_path)

    # Set Variables for new Workbook and Worksheets
    new_workbook = load_workbook(new_file_path)
    new_workbook_pay_1_worksheet = new_workbook["{0} - Pay 1".format(new_month_abbr)]
    new_workbook_pay_2_worksheet = new_workbook["{0} - Pay 2".format(new_month_abbr)]
    new_workbook_pay_3_worksheet = new_workbook["{0} - Pay 3".format(new_month_abbr)]
    new_workbook_receipt_1_worksheet = new_workbook["Receipt 1"]
    new_workbook_receipt_2_worksheet = new_workbook["Receipt 2"]
    new_workbook_receipt_3_worksheet = new_workbook["Receipt 3"]

    # Set static variables for biweekly finance excel cell, column, and row values
    pay_start_date_cell = "B25"
    pay_end_date_cell = "B26"
    receipt_start_date_cell = "B24"
    receipt_end_date_cell = "B25"
    total_amount_column_letter = "C"
    new_total_amount_column_letter = "H"
    money_category_start_row_inclusive = 2
    money_category_end_row_exclusive = 22

    # Pulling Values
    if os.path.isfile(previous_file_path):  # Previous Month's workbook is found
        previous_workbook = load_workbook(previous_file_path, data_only=True)
        previous_workbook_datetime = _determine_biweekly_finance_file_datetime(previous_file_path)
        previous_month_abbr = previous_workbook_datetime.strftime('%b')

        # Previous Workbook most recent worksheet
        print("Determining most recent pay worksheet...")
        if "{0} - Pay 3".format(previous_month_abbr) in previous_workbook.sheetnames:
            # Use Pay 3
            previous_workbook_most_recent_pay_worksheet = previous_workbook["{0} - Pay 3".format(previous_month_abbr)]
        elif "{0} - Pay 2".format(previous_month_abbr) in previous_workbook.sheetnames:
            # Use Pay 2
            previous_workbook_most_recent_pay_worksheet = previous_workbook["{0} - Pay 2".format(previous_month_abbr)]
        else:
            # Use Pay 1
            previous_workbook_most_recent_pay_worksheet = previous_workbook["{0} - Pay 1".format(previous_month_abbr)]
        previous_workbook_last_date = previous_workbook_most_recent_pay_worksheet[pay_end_date_cell].value

        # Paste old workbook "New Total Amount" in new workbook's "Total Amount" for each category
        print("Calculating category money values...")
        for row in range(money_category_start_row_inclusive, money_category_end_row_exclusive):
            src_cell = new_total_amount_column_letter + str(row)
            dst_cell = total_amount_column_letter + str(row)
            value = previous_workbook_most_recent_pay_worksheet[src_cell].value
            value2 = previous_workbook_most_recent_pay_worksheet[src_cell]
            new_workbook_pay_1_worksheet[dst_cell].value = value
        new_workbook.save(new_file_path)

    else:  # No previous file
        previous_workbook_datetime = _determine_biweekly_finance_file_datetime(previous_file_path)
        first_day = int(input("No workbook found for previous month to determine start date of the new workbook. "
                              "\nInput start day of this workbook: "))
        first_month = new_workbook_datetime.month
        first_year = new_workbook_datetime.year
        new_workbook_first_date = datetime(month=first_month, year=first_year, day=first_day)
        previous_workbook_last_date = new_workbook_first_date - timedelta(days=1)

    # Calculate Dates
    print("Calculating dates for new worksheets...")
    new_biweekly_1_start_date = previous_workbook_last_date + timedelta(days=1)
    new_biweekly_1_end_date = new_biweekly_1_start_date + timedelta(days=13)
    new_biweekly_2_start_date = new_biweekly_1_end_date + timedelta(days=1)
    new_biweekly_2_end_date = new_biweekly_2_start_date + timedelta(days=13)
    new_biweekly_3_start_date = new_biweekly_2_end_date + timedelta(days=1)
    new_biweekly_3_end_date = new_biweekly_3_start_date + timedelta(days=13)

    # Paste new workbook dates
    new_workbook_pay_1_worksheet[pay_start_date_cell] = new_biweekly_1_start_date
    new_workbook_pay_1_worksheet[pay_end_date_cell] = new_biweekly_1_end_date
    new_workbook_receipt_1_worksheet[receipt_start_date_cell] = new_biweekly_1_start_date
    new_workbook_receipt_1_worksheet[receipt_end_date_cell] = new_biweekly_1_end_date
    new_workbook_pay_2_worksheet[pay_start_date_cell] = new_biweekly_2_start_date
    new_workbook_pay_2_worksheet[pay_end_date_cell] = new_biweekly_2_end_date
    new_workbook_receipt_2_worksheet[receipt_start_date_cell] = new_biweekly_2_start_date
    new_workbook_receipt_2_worksheet[receipt_end_date_cell] = new_biweekly_2_end_date
    new_workbook_pay_3_worksheet[pay_start_date_cell] = new_biweekly_3_start_date
    new_workbook_pay_3_worksheet[pay_end_date_cell] = new_biweekly_3_end_date
    new_workbook_receipt_3_worksheet[receipt_start_date_cell] = new_biweekly_3_start_date
    new_workbook_receipt_3_worksheet[receipt_end_date_cell] = new_biweekly_3_end_date
    new_workbook.save(new_file_path)

    # Update "Total Amount" formulas to point to previous pay in same workbook instead of template workbook
    for row in range(money_category_start_row_inclusive, money_category_end_row_exclusive):
        cell = total_amount_column_letter + str(row)
        pay_2_cell_value = new_workbook_pay_2_worksheet[cell].value.replace('Pay 1', f'{new_month_abbr} - Pay 1')
        pay_3_cell_value = new_workbook_pay_3_worksheet[cell].value.replace('Pay 2', f'{new_month_abbr} - Pay 2')
        new_workbook_pay_2_worksheet[cell].value = pay_2_cell_value
        new_workbook_pay_3_worksheet[cell].value = pay_3_cell_value
    new_workbook.save(new_file_path)

    # Remove Pay 3 and Receipt 3 if it should not be in the new workbook (if end date is into next month)
    _, num_of_days_in_month = calendar.monthrange(new_workbook_datetime.year, new_workbook_datetime.month)
    last_day_of_month = datetime(new_workbook_datetime.year, new_workbook_datetime.month, num_of_days_in_month)
    if new_biweekly_3_end_date > last_day_of_month:
        print("3rd pay end date is into next month. Removing 3rd Pay from this worksheet...")
        new_workbook.remove(new_workbook_pay_3_worksheet)
        new_workbook.remove(new_workbook_receipt_3_worksheet)
        new_workbook.save(new_file_path)
    if new_biweekly_2_end_date > last_day_of_month:
        print("2nd pay end date is into next month. Removing 2nd Pay from this worksheet...")
        new_workbook.remove(new_workbook_pay_2_worksheet)
        new_workbook.remove(new_workbook_receipt_2_worksheet)
        new_workbook.save(new_file_path)


def fetch_biweekly_finance_file_path(file_date: datetime, biweekly_finance_dir: str = sys.path[0]):
    """
    Returns the biweekly finance file of the specified date
    :param file_date: date of the biweekly finance file
    :param biweekly_finance_dir: base directory of all biweekly finance files
    :return: filepath of biweekly finance file
    :rtype: str
    """
    year = file_date.strftime('%Y')
    month_str = file_date.strftime('%B')
    return os.path.join(biweekly_finance_dir, year, f'{year} - {month_str}.xlsx')


def _copy_and_rename_file(file_path: str, dest_dir: str, new_file_name):
    """
    Copies the specified file to the dest_dir (creating the directory if necessary) and renames it to the new_file_name
    :param file_path: file path of the file to copy
    :param dest_dir: directory to copy the file to
    :param new_file_name: name the file should be changed to
    :return: file path of the new file
    """
    # Copy File
    try:
        # Creating new directory with year if does not exist
        os.makedirs(dest_dir, exist_ok=True)
        # Copying File
        print("Copying file: {0}".format(file_path))
        # new_file_copy = shutil.copyfile(file_path, dest_dir)
        new_file_copy = shutil.copy(file_path, dest_dir)
        print("Copied file to {0}".format(dest_dir))
        # Renaming File
        print("Renaming file: {0}".format(new_file_copy))
        new_file_path = os.path.join(dest_dir, new_file_name)
        os.rename(src=new_file_copy, dst=new_file_path)
        print("File successfully renamed to " + new_file_path)
        return new_file_path
    except Exception as e:
        print("Failed to copy or rename file.")
        print(e)


def _determine_biweekly_finance_file_path(file_date: datetime, biweekly_finance_dir: str = sys.path[0]):
    """
    Returns the file path of the biweekly finance .xlsx file of the given file date month and year
    :param file_date: str month and year of the biweekly finance .xlsx file
    :param biweekly_finance_dir: str file path of the directory where all biweekly finance files are located
    :return: file path
    :rtype: str
    """
    year = file_date.strftime('%Y')
    month_str = file_date.strftime('%B')
    return os.path.join(biweekly_finance_dir, year, f'{year} - {month_str}.xlsx')


def _determine_biweekly_finance_file_datetime(file_path: str):
    """
    Returns a datetime object of the biweekly finance file month and year
    :param file_path: biweekly finance .xlsx file
    :return: datetime month and year of the file
    :rtype: datetime
    """
    file_name = os.path.basename(file_path)
    return datetime.strptime(file_name, '%Y - %B.xlsx')


###############################################################
# Main Method
###############################################################
if __name__ == "__main__":
    # Initialize Static Variables
    TEMPLATE_FILE_PATH = os.path.join(sys.path[0], 'Biweekly Finances Template', 'biweekly_finances_template.xlsx')
    BIWEEKLY_FINANCES_DIR_PATH = sys.path[0]

    # Input
    input_month_num = int(input("Month Number of new Excel Sheet: "))
    input_year = int(input("Year of new Excel Sheet: "))

    # Determine datetime of new file date and previous file date based on input
    new_file_date = datetime(year=input_year, month=input_month_num, day=1)
    previous_file_date = new_file_date - relativedelta.relativedelta(months=1)

    # Create new .xlsx file from template
    new_finance_file_path = create_new_biweekly_finance_file(file_date=new_file_date,
                                                             template_file_path=TEMPLATE_FILE_PATH)

    # Get previous month's biweekly finance .xlsx file
    previous_finance_file_path = fetch_biweekly_finance_file_path(file_date=previous_file_date,
                                                                  biweekly_finance_dir=BIWEEKLY_FINANCES_DIR_PATH)

    # Determine new values based on previous month's Biweekly finance .xlsx file
    populate_biweekly_finance_values(new_file_path=new_finance_file_path, previous_file_path=previous_finance_file_path)

    # Wait to exit
    print("")
    input("PRESS ENTER TO EXIT...")
